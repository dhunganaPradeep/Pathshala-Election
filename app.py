from flask import Flask, render_template, redirect, url_for, request, session, flash, jsonify, send_file, g
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import os
import sqlite3
import secrets
import string
import pandas as pd
from datetime import datetime, timedelta
import io
from fpdf import FPDF
import tempfile
import openpyxl
from flask_wtf.csrf import CSRFProtect
from flask_session import Session
import logging
import time
import re

app = Flask(__name__)

# Use a strong random secret key - in production, set this via environment variable
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx'}
app.config['DATABASE'] = os.environ.get('DATABASE_PATH', 'election.db')
app.config['ADMIN_TIMEOUT'] = int(os.environ.get('ADMIN_TIMEOUT', 1800))
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5MB limit

# Session security settings
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_FILE_DIR'] = os.environ.get('SESSION_FILE_DIR', 'flask_session')
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_USE_SIGNER'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)
app.config['SESSION_COOKIE_SECURE'] = True  # Ensure cookies only sent over HTTPS
app.config['SESSION_COOKIE_HTTPONLY'] = True  # Prevent JavaScript from accessing cookies
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # Protect against CSRF attacks

# Configure logging with minimal output
log_level = os.environ.get('LOG_LEVEL', 'ERROR')
log_file = os.environ.get('LOG_FILE', 'app.log')
logging.basicConfig(level=getattr(logging, log_level), 
                    format='%(asctime)s %(levelname)s: %(message)s',
                    handlers=[logging.FileHandler(log_file)])

# Initialize CSRF protection
csrf = CSRFProtect(app)
Session(app)

# Exempt routes that use JSON from CSRF protection
csrf.exempt('/verify_code')
csrf.exempt('/cast_vote')

# Set security headers for all responses
@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self' 'unsafe-inline' https://cdn.tailwindcss.com https://code.jquery.com https://cdnjs.cloudflare.com; style-src 'self' 'unsafe-inline' https://cdnjs.cloudflare.com; img-src 'self' data:; font-src 'self' https://cdnjs.cloudflare.com;"
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    
    # Add cache control headers to prevent back button issues
    if request.path.startswith('/admin'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '-1'
    
    return response

# Ensure necessary directories exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['SESSION_FILE_DIR'], exist_ok=True)
os.makedirs(os.path.join('static', 'img', 'candidates'), exist_ok=True)
os.makedirs(os.path.join('static', 'img', 'logos'), exist_ok=True)

# SQL Injection prevention - parameterized query wrapper
def execute_safe_query(conn, query, params=None):
    """Execute a parameterized query to prevent SQL injection"""
    if params is None:
        params = ()
    return conn.execute(query, params)

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(app.config['DATABASE'])
        g.db.row_factory = sqlite3.Row
    return g.db

@app.teardown_appcontext
def close_db(error):
    db = g.pop('db', None)
    if db is not None:
        db.close()

def get_db_connection():
    return get_db()

class DBContextManager:
    def __enter__(self):
        self.conn = get_db()
        return self.conn
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type:
            logging.error(f"Database error: {str(exc_val)}")
            return False
        self.conn.commit()
        return False

# Input validation function
def sanitize_input(input_string, pattern=r'[^a-zA-Z0-9_\-\s]', replacement=''):
    """Sanitize input by removing potentially dangerous characters"""
    if input_string:
        return re.sub(pattern, replacement, input_string)
    return input_string

def init_db():
    try:
        logging.info("Initializing database...")
        with app.app_context():
            conn = get_db()
            with open('schema.sql') as f:
                conn.executescript(f.read())
            conn.commit()
            logging.info("Database initialized successfully")
    except sqlite3.Error as e:
        logging.error(f"Database initialization error: {str(e)}")
        raise
    except Exception as e:
        logging.error(f"Error initializing database: {str(e)}")
        raise

def generate_unique_code(length=6):
    chars = string.ascii_uppercase + string.digits
    max_attempts = 10  
    
    for attempt in range(max_attempts):
        code = ''.join(secrets.choice(chars) for _ in range(length)).upper()
        
        try:
            conn = get_db()
            result = conn.execute(
                'SELECT * FROM voters WHERE voting_code = ?', 
                (code,)
            ).fetchone()
            
            if result is None:
                return code
        except sqlite3.Error as e:
            logging.error(f"Database error in generate_unique_code: {str(e)}")
            code = ''.join(secrets.choice(chars) for _ in range(length + 1)).upper()
            return code
    
    return ''.join(secrets.choice(chars) for _ in range(length + 1)).upper()

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

@app.before_request
def check_admin_default_password():
    if request.endpoint and request.endpoint.startswith('admin_') and request.endpoint not in ['admin_login', 'admin_logout', 'admin_setup']:
        if session.get('admin'):
            if session.get('default_password'):
                flash("Security warning: Your account is using an insecure password. You must change it before proceeding.", "warning")
                return redirect(url_for('admin_setup'))
            
            with DBContextManager() as conn:
                admin = conn.execute('SELECT id, username, password_hash FROM admin WHERE id = ?', (session.get('admin_id'),)).fetchone()
                
                if admin:
                    admin_variants = ['admin', 'Admin', 'ADMIN', 'administrator', 'admin123']
                    is_using_admin_password = False
                    
                    for variant in admin_variants:
                        if check_password_hash(admin['password_hash'], variant):
                            is_using_admin_password = True
                            break
                    
                    if admin['password_hash'] == '$2b$12$qKU3YP7Nz3kzWkxpYKiMqe4JfN9aKC7GW4q1Eb1iiL6TgW/LQTKCm':
                        is_using_admin_password = True
                    
                    if is_using_admin_password:
                        session['default_password'] = True
                        flash("Security warning: Your account is using an insecure password. You must change it before proceeding.", "warning")
                        return redirect(url_for('admin_setup'))
    
@app.before_request
def check_admin_timeout():
    if 'admin' in session and 'admin_last_activity' in session:
        last_activity = session.get('admin_last_activity')
        current_time = datetime.now()
        
        try:
            if isinstance(last_activity, str):
                last_activity = datetime.fromisoformat(last_activity)
            elif not isinstance(last_activity, datetime):
                session['admin_last_activity'] = current_time.isoformat()
                return

            if hasattr(last_activity, 'tzinfo') and last_activity.tzinfo is not None:
                last_activity = last_activity.replace(tzinfo=None)
                
            if current_time - last_activity > timedelta(seconds=app.config['ADMIN_TIMEOUT']):
                session.pop('admin', None)
                session.pop('admin_id', None)
                session.pop('admin_username', None)
                session.pop('admin_last_activity', None)
                if request.endpoint and request.endpoint.startswith('admin_'):
                    flash('Your session has expired due to inactivity')
                    return redirect(url_for('admin_login'))
        except (ValueError, TypeError):
            session['admin_last_activity'] = current_time.isoformat()
            return
    
    if 'admin' in session:
        session['admin_last_activity'] = datetime.now().isoformat()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/verify_code', methods=['POST'])
def verify_code():
    try:
        # Get and clean the code
        code = request.form.get('code', '').strip().upper()
        code = code.replace(' ', '')
        
        logging.info(f"Verifying code: {code}")
        
        if not code:
            return jsonify({'success': False, 'message': 'Please enter a code'})
        
        # Validate code format
        if not code.isalnum() or len(code) > 8:
            logging.warning(f"Invalid code format: {code}")
            return jsonify({'success': False, 'message': 'Invalid code format'})
        
        conn = get_db()
        voter = conn.execute('SELECT * FROM voters WHERE UPPER(voting_code) = ?', (code,)).fetchone()
        
        # Check if voter exists
        if voter is None:
            logging.info(f"Failed verification attempt with code: {code}")
            return jsonify({'success': False, 'message': 'Invalid code'})
        
        # Check if already voted
        if voter['has_voted']:
            logging.info(f"Voter {voter['id']} with code {code} attempted to vote again")
            return jsonify({'success': False, 'message': 'You cannot vote now since it is already voted'})
        
        # Set session variables
        session['voter_id'] = voter['id']
        session['is_teacher'] = bool(voter['is_teacher'])
        session.modified = True 
        
        # Normalize the voting code if needed
        if voter['voting_code'] != code:
            logging.info(f"Normalizing voting code for voter {voter['id']} from {voter['voting_code']} to {code}")
            conn.execute('UPDATE voters SET voting_code = ? WHERE id = ?', (code, voter['id']))
            conn.commit()
        
        logging.info(f"Voter {voter['id']} successfully verified with code {code}, is_teacher: {bool(voter['is_teacher'])}")
        return jsonify({'success': True})
    
    except sqlite3.Error as e:
        logging.error(f"Database error in verify_code: {str(e)}")
        return jsonify({'success': False, 'message': 'A database error occurred. Please try again.'})
    except Exception as e:
        logging.error(f"Error in verify_code: {str(e)}, Type: {type(e).__name__}")
        return jsonify({'success': False, 'message': 'An error occurred. Please try again.'})

@app.route('/vote')
def vote():
    if 'voter_id' not in session:
        logging.warning("Attempted to access vote page without voter_id in session")
        flash("Please enter your voting code first", "error")
        return redirect(url_for('index'))
    
    try:
        voter_id = session.get('voter_id')
        logging.info(f"Voter {voter_id} accessing vote page")
        
        conn = get_db()
        voter = conn.execute('SELECT * FROM voters WHERE id = ? AND has_voted = 0', (voter_id,)).fetchone()
        
        if not voter:
            logging.warning(f"Voter {voter_id} not found or has already voted")
            session.pop('voter_id', None)
            session.pop('is_teacher', None)
            session.modified = True
            flash("Your voting session is no longer valid", "error")
            return redirect(url_for('index'))
        
        query = '''
            SELECT id, name, gender, image_path, logo_path, slogan, description
            FROM candidates 
            WHERE gender = 'Male'
            UNION ALL
            SELECT id, name, gender, image_path, logo_path, slogan, description
            FROM candidates 
            WHERE gender = 'Female'
        '''
        candidates = conn.execute(query).fetchall()
        
        male_candidates = []
        female_candidates = []
        
        for candidate in candidates:
            candidate_dict = dict(candidate)
            if candidate_dict['gender'] == 'Male':
                male_candidates.append(candidate_dict)
            else:
                female_candidates.append(candidate_dict)
        
        if not male_candidates or not female_candidates:
            logging.warning("No candidates available for voting")
            flash('There are no candidates available for voting at this time.')
            session.pop('voter_id', None)
            session.pop('is_teacher', None)
            session.modified = True
            return redirect(url_for('index'))
        
        return render_template('vote.html', 
                           male_candidates=male_candidates, 
                           female_candidates=female_candidates,
                           is_teacher=session.get('is_teacher', False))
    except sqlite3.Error as e:
        logging.error(f"Database error in vote page: {str(e)}")
        flash('An error occurred while loading candidates. Please try again.')
        return redirect(url_for('index'))
    except Exception as e:
        logging.error(f"Error in vote page: {str(e)}")
        flash('An unexpected error occurred. Please try again.')
        return redirect(url_for('index'))

@app.route('/cast_vote', methods=['POST'])
def cast_vote():
    if 'voter_id' not in session:
        return jsonify({'success': False, 'message': 'You are not authorized to vote. Please enter your voting code first.'})
    
    voter_id = session.get('voter_id')
    
    try:
        data = request.get_json()
        if not data:
            logging.error(f"No JSON data received in cast_vote. Content-Type: {request.headers.get('Content-Type')}")
            return jsonify({'success': False, 'message': 'Invalid request format. Please try again.'})
        
        male_candidate_id = data.get('male_candidate_id')
        female_candidate_id = data.get('female_candidate_id')
        
        if not male_candidate_id or not female_candidate_id:
            logging.error(f"Missing candidate ID in cast_vote: male={male_candidate_id}, female={female_candidate_id}")
            return jsonify({'success': False, 'message': 'You must select both a male and female candidate.'})
        
        with DBContextManager() as conn:
            conn.execute('BEGIN EXCLUSIVE TRANSACTION')
            
            voter = conn.execute('SELECT has_voted FROM voters WHERE id = ? AND has_voted = 0', (voter_id,)).fetchone()
            
            if not voter:
                conn.rollback()
                logging.warning(f"Voter {voter_id} attempted to vote but has already voted or is invalid")
                return jsonify({'success': False, 'message': 'You have already voted or your voter ID is invalid.'})
            
            candidates = conn.execute('''
                SELECT 
                    SUM(CASE WHEN id = ? AND gender = 'Male' THEN 1 ELSE 0 END) as valid_male,
                    SUM(CASE WHEN id = ? AND gender = 'Female' THEN 1 ELSE 0 END) as valid_female
                FROM candidates
                WHERE id IN (?, ?)
            ''', (male_candidate_id, female_candidate_id, male_candidate_id, female_candidate_id)).fetchone()
            
            if not candidates['valid_male']:
                conn.rollback()
                logging.warning(f"Invalid male candidate ID: {male_candidate_id}")
                return jsonify({'success': False, 'message': 'Invalid male candidate selection.'})
            
            if not candidates['valid_female']:
                conn.rollback()
                logging.warning(f"Invalid female candidate ID: {female_candidate_id}")
                return jsonify({'success': False, 'message': 'Invalid female candidate selection.'})
            
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            try:
                conn.executemany(
                    'INSERT INTO votes (voter_id, candidate_id, timestamp) VALUES (?, ?, ?)',
                    [(voter_id, male_candidate_id, timestamp), (voter_id, female_candidate_id, timestamp)]
                )
                
                conn.execute('UPDATE voters SET has_voted = 1 WHERE id = ?', (voter_id,))
                conn.commit()
                
                session.pop('voter_id', None)
                session.pop('is_teacher', None)
                session.modified = True
                
                logging.info(f"Voter {voter_id} successfully cast votes for candidates {male_candidate_id} and {female_candidate_id}")
                return jsonify({'success': True})
                
            except sqlite3.Error as e:
                conn.rollback()
                logging.error(f"Database error in vote casting: {str(e)}")
                return jsonify({'success': False, 'message': f'A database error occurred: {str(e)}'})
                
    except sqlite3.Error as e:
        logging.error(f"Database error in cast_vote: {str(e)}")
        return jsonify({'success': False, 'message': f'A database error occurred: {str(e)}'})
    except Exception as e:
        logging.error(f"Error in cast_vote: {str(e)}, Type: {type(e).__name__}")
        return jsonify({'success': False, 'message': f'An error occurred: {str(e)}'})

# Admin routes
@app.route('/admin')
def admin_login():
    if session.get('admin'):
        with DBContextManager() as conn:
            admin = conn.execute('SELECT password_hash FROM admin WHERE id = ?', (session['admin_id'],)).fetchone()
            
            if admin:
                admin_variants = ['admin', 'Admin', 'ADMIN', 'administrator', 'admin123']
                is_using_admin_password = False
                
                for variant in admin_variants:
                    if check_password_hash(admin['password_hash'], variant):
                        is_using_admin_password = True
                        break
                
                if admin['password_hash'] == '$2b$12$qKU3YP7Nz3kzWkxpYKiMqe4JfN9aKC7GW4q1Eb1iiL6TgW/LQTKCm':
                    is_using_admin_password = True
                
                if is_using_admin_password:
                    session['default_password'] = True
                    flash("Security warning: Your account is using an insecure password. You must change it before proceeding.", "warning")
                    return redirect(url_for('admin_setup'))
        return redirect(url_for('admin_dashboard'))
    return render_template('admin/login.html')

@app.route('/admin/login', methods=['POST'])
def handle_admin_login():
    # Apply rate limiting for login attempts
    ip_address = request.remote_addr
    current_time = time.time()
    
    # Get login attempts from the session
    login_attempts = session.get('login_attempts', {})
    
    # Clean up old attempts (older than 30 minutes)
    login_attempts = {ip: data for ip, data in login_attempts.items() 
                     if current_time - data['timestamp'] < 1800}
    
    # Check if IP is already blocked
    if ip_address in login_attempts and login_attempts[ip_address]['count'] >= 5:
        if current_time - login_attempts[ip_address]['timestamp'] < 1800:  # 30 minutes
            flash('Too many failed login attempts. Please try again later.', 'error')
            session['login_attempts'] = login_attempts
            return redirect(url_for('admin_login'))
    
    # Get and sanitize credentials
    username = sanitize_input(request.form.get('username', '').strip())
    password = request.form.get('password', '')
    
    # For debugging - log admin login attempts without password
    logging.info(f"Admin login attempt: username={username}")
    
    if not username or not password:
        flash('Username and password are required', 'error')
        return redirect(url_for('admin_login'))
    
    try:
        with DBContextManager() as conn:
            # Use parameterized query to prevent SQL injection
            admin = execute_safe_query(
                conn, 
                'SELECT id, username, password_hash FROM admin WHERE username = ?', 
                (username,)
            ).fetchone()
            
            # No admin found with this username
            if not admin:
                # Record failed login attempt
                if ip_address in login_attempts:
                    login_attempts[ip_address]['count'] += 1
                    login_attempts[ip_address]['timestamp'] = current_time
                else:
                    login_attempts[ip_address] = {'count': 1, 'timestamp': current_time}
                
                session['login_attempts'] = login_attempts
                
                # Log failed attempt
                logging.warning(f"Failed login attempt for nonexistent username: {username} from IP: {ip_address}")
                flash('Invalid username or password', 'error')
                return redirect(url_for('admin_login'))
                
            # For debugging
            logging.info(f"Admin found with ID: {admin['id']}, checking password...")
            
            # Default admin account with default password
            if admin['password_hash'] == '$2b$12$qKU3YP7Nz3kzWkxpYKiMqe4JfN9aKC7GW4q1Eb1iiL6TgW/LQTKCm' and username == 'admin' and password == 'admin':
                logging.info(f"Default admin password login success for: {username}")
                
                # Reset login attempts for this IP
                if ip_address in login_attempts:
                    login_attempts.pop(ip_address)
                session['login_attempts'] = login_attempts
                
                # Set admin session
                session['admin'] = True
                session['admin_id'] = admin['id']
                session['admin_username'] = admin['username']
                session['admin_last_activity'] = datetime.now().isoformat()
                session['default_password'] = True
                
                # Log successful login
                logging.info(f"Admin {username} logged in with default password")
                
                flash("You must change the default admin password before continuing", "warning")
                return redirect(url_for('admin_setup'))
            
            # Check password with standard hash comparison
            # Try with and without whitespace in case there are invisible characters
            password_correct = check_password_hash(admin['password_hash'], password)
            password_trimmed_correct = check_password_hash(admin['password_hash'], password.strip())
            
            # Valid credentials with proper password
            if password_correct or password_trimmed_correct:
                logging.info(f"Password verification successful for admin: {username}")
                
                # Reset login attempts for this IP
                if ip_address in login_attempts:
                    login_attempts.pop(ip_address)
                session['login_attempts'] = login_attempts
                
                # Set admin session
                session['admin'] = True
                session['admin_id'] = admin['id']
                session['admin_username'] = admin['username']
                session['admin_last_activity'] = datetime.now().isoformat()
                
                # Log successful login
                logging.info(f"Admin {username} logged in successfully")
                
                # Check for weak passwords
                admin_variants = ['admin', 'Admin', 'ADMIN', 'administrator', 'admin123']
                if password in admin_variants:
                    session['default_password'] = True
                    flash("Security warning: Your account is using an insecure password. You must change it before proceeding.", "warning")
                    return redirect(url_for('admin_setup'))
                
                return redirect(url_for('admin_dashboard'))
            
            # Invalid credentials
            else:
                logging.warning(f"Password verification failed for admin: {username}")
                
                # Record failed login attempt
                if ip_address in login_attempts:
                    login_attempts[ip_address]['count'] += 1
                    login_attempts[ip_address]['timestamp'] = current_time
                else:
                    login_attempts[ip_address] = {'count': 1, 'timestamp': current_time}
                
                session['login_attempts'] = login_attempts
                
                # Log failed attempt
                logging.warning(f"Failed login attempt for username: {username} from IP: {ip_address} (Attempt #{login_attempts[ip_address]['count']})")
                
                if login_attempts[ip_address]['count'] >= 5:
                    flash('Too many failed login attempts. Please try again later.', 'error')
                else:
                    flash('Invalid username or password', 'error')
                
                return redirect(url_for('admin_login'))
                
    except Exception as e:
        logging.error(f"Error during login: {str(e)}")
        flash('An error occurred during login. Please try again.', 'error')
        return redirect(url_for('admin_login'))

@app.route('/admin/setup', methods=['GET', 'POST'])
def admin_setup():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    
    with DBContextManager() as conn:
        admin = conn.execute('SELECT password_hash FROM admin WHERE id = ?', (session['admin_id'],)).fetchone()
        
        if admin:
            admin_variants = ['admin', 'Admin', 'ADMIN', 'administrator', 'admin123']
            is_using_insecure_password = False
            
            for variant in admin_variants:
                if check_password_hash(admin['password_hash'], variant):
                    is_using_insecure_password = True
                    break
            
            if admin['password_hash'] == '$2b$12$qKU3YP7Nz3kzWkxpYKiMqe4JfN9aKC7GW4q1Eb1iiL6TgW/LQTKCm':
                is_using_insecure_password = True
            
            if not is_using_insecure_password and not session.get('default_password'):
                if 'default_password' in session:
                    session.pop('default_password', None)
                return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        if not new_password or not confirm_password:
            flash('Both fields are required', 'error')
            return render_template('admin/setup.html')
        
        if new_password != confirm_password:
            flash('Passwords do not match', 'error')
            return render_template('admin/setup.html')
        
        admin_variants = ['admin', 'Admin', 'ADMIN', 'administrator', 'admin123']
        if any(variant in new_password.lower() for variant in ['admin', 'administrator']):
            flash('Your password cannot contain "admin" or "administrator"', 'error')
            return render_template('admin/setup.html')
        
        if len(new_password) < 8:
            flash('Password must be at least 8 characters long', 'error')
            return render_template('admin/setup.html')
        
        try:
            with DBContextManager() as conn:
                new_password_hash = generate_password_hash(new_password)
                conn.execute('UPDATE admin SET password_hash = ? WHERE id = ?', 
                          (new_password_hash, session['admin_id']))
                conn.commit()
                
                logging.info(f"Admin {session['admin_username']} successfully changed password")
                
                if 'default_password' in session:
                    session.pop('default_password', None)
                
                flash('Password successfully changed', 'success')
                return redirect(url_for('admin_dashboard'))
        except Exception as e:
            logging.error(f"Error changing password: {str(e)}")
            flash(f'Error changing password: {str(e)}', 'error')
            return render_template('admin/setup.html')
    
    return render_template('admin/setup.html')

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin', None)
    session.pop('admin_id', None)
    session.pop('admin_username', None)
    return redirect(url_for('admin_login'))

@app.route('/admin/dashboard')
def admin_dashboard():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    
    try:
        with DBContextManager() as conn:
            # Consolidated query to get all voter statistics in one go
            stats = conn.execute('''
                SELECT 
                    SUM(CASE WHEN is_teacher = 0 THEN 1 ELSE 0 END) as total_students,
                    SUM(CASE WHEN is_teacher = 1 THEN 1 ELSE 0 END) as total_teachers,
                    SUM(CASE WHEN has_voted = 1 THEN 1 ELSE 0 END) as total_votes,
                    SUM(CASE WHEN is_teacher = 0 AND has_voted = 1 THEN 1 ELSE 0 END) as students_voted,
                    SUM(CASE WHEN is_teacher = 1 AND has_voted = 1 THEN 1 ELSE 0 END) as teachers_voted
                FROM voters
            ''').fetchone()
            
            total_students = stats['total_students'] or 0
            total_teachers = stats['total_teachers'] or 0
            total_votes = stats['total_votes'] or 0
            students_voted = stats['students_voted'] or 0
            teachers_voted = stats['teachers_voted'] or 0
            votes_remaining = total_students + total_teachers - total_votes
            
            student_vote_percent = (students_voted / total_students * 100) if total_students > 0 else 0
            teacher_vote_percent = (teachers_voted / total_teachers * 100) if total_teachers > 0 else 0
            
            # Fetch students with a proper LIMIT for pagination if needed
            students = conn.execute('SELECT * FROM voters WHERE is_teacher = 0 LIMIT 100').fetchall()
            
            return render_template('admin/dashboard.html',
                                total_students=total_students,
                                total_teachers=total_teachers,
                                total_votes=total_votes,
                                votes_remaining=votes_remaining,
                                students=students,
                                students_voted=students_voted,
                                teachers_voted=teachers_voted,
                                student_vote_percent=student_vote_percent,
                                teacher_vote_percent=teacher_vote_percent)
    except Exception as e:
        logging.error(f"Error loading dashboard: {str(e)}")
        flash(f'Error loading dashboard: {str(e)}')
        return redirect(url_for('admin_login'))

# Student Management Routes
@app.route('/admin/students')
def admin_students():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    
    try:
        with DBContextManager() as conn:
            students = conn.execute('SELECT * FROM voters WHERE is_teacher = 0').fetchall()
            return render_template('admin/students.html', students=students)
    except Exception as e:
        flash(f'Error loading students: {str(e)}')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/students/upload', methods=['POST'])
def upload_students():
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file part'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'}), 400
    
    if not file.filename.lower().endswith('.xlsx'):
        return jsonify({'success': False, 'message': 'File must be an Excel (.xlsx) file'}), 400
    
    # Check file size (limit to 5MB)
    file_size = file.content_length or 0
    if file_size > 5 * 1024 * 1024:  # 5MB
        return jsonify({'success': False, 'message': 'File size exceeds the 5MB limit'}), 400
    
    temp_file = None
    pdf_output = None
    
    try:
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        temp_file.close()
        
        try:
            workbook = openpyxl.load_workbook(temp_file.name, read_only=True, data_only=True)
            sheet = workbook.active
            
            if sheet.max_row < 2:  
                os.unlink(temp_file.name) 
                return jsonify({'success': False, 'message': 'Excel file is empty or has no data rows'}), 400
            
            if sheet.max_row > 1000:  
                os.unlink(temp_file.name) 
                return jsonify({'success': False, 'message': 'Excel file has too many rows (limit: 1000)'}), 400
            
            header_row = [cell.value for cell in sheet[1]]
            
            required_columns = ['Name', 'Class', 'Section', 'Roll No']
            for col in required_columns:
                if col not in header_row:
                    os.unlink(temp_file.name)  
                    return jsonify({
                        'success': False, 
                        'message': f'Missing required column: {col}. Please make sure your Excel file has the columns: {", ".join(required_columns)}'
                    }), 400
            
            name_col = header_row.index('Name') + 1
            class_col = header_row.index('Class') + 1
            section_col = header_row.index('Section') + 1
            roll_col = header_row.index('Roll No') + 1
            
            students_added = 0
            students_updated = 0
            processed_students = set()
            
            with DBContextManager() as conn:
                pdf = FPDF()
                pdf.add_page()
                pdf.set_font("Arial", size=12)
                
                pdf.cell(200, 10, txt="Pathshala Election 2082 - Student Voting Codes", ln=True, align='C')
                pdf.set_font("Arial", 'B', size=14)
                pdf.cell(200, 10, txt="CONFIDENTIAL", ln=True, align='C')
                pdf.set_font("Arial", size=10)
                pdf.cell(200, 10, txt=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align='C')
                pdf.ln(10)
                
                pdf.set_font("Arial", 'B', size=10)
                pdf.cell(60, 10, txt="Name", border=1)
                pdf.cell(20, 10, txt="Class", border=1)
                pdf.cell(20, 10, txt="Section", border=1)
                pdf.cell(20, 10, txt="Roll No", border=1)
                pdf.cell(30, 10, txt="Voting Code", border=1, ln=True)
                
                pdf.set_font("Arial", size=10)
                
                for row in range(2, sheet.max_row + 1):
                    name = sheet.cell(row=row, column=name_col).value
                    class_val = sheet.cell(row=row, column=class_col).value
                    section = sheet.cell(row=row, column=section_col).value
                    roll_no = sheet.cell(row=row, column=roll_col).value
                    
                    if not name or not class_val or not section:
                        continue
                    
                    name = str(name)[:100].strip() 
                    class_val = str(class_val)[:20].strip()
                    section = str(section)[:20].strip()
                    roll_no = str(roll_no) if roll_no is not None else ''
                    roll_no = roll_no[:20].strip()
                    
                    student_key = f"{class_val}_{section}_{roll_no}"
                    if student_key in processed_students:
                        continue
                    
                    processed_students.add(student_key)
                    
                    voting_code = generate_unique_code()
                    
                    existing = conn.execute(
                        'SELECT id, has_voted FROM voters WHERE class = ? AND section = ? AND roll_no = ? AND is_teacher = 0',
                        (class_val, section, roll_no)
                    ).fetchone()
                    
                    if existing:
                        conn.execute(
                            'UPDATE voters SET voting_code = ?, name = ? WHERE id = ?',
                            (voting_code, name, existing['id'])
                        )
                        students_updated += 1
                    else:
                        conn.execute(
                            'INSERT INTO voters (name, class, section, roll_no, voting_code, is_teacher) VALUES (?, ?, ?, ?, ?, 0)',
                            (name, class_val, section, roll_no, voting_code)
                        )
                        students_added += 1
                    
                    pdf.cell(60, 10, txt=name, border=1)
                    pdf.cell(20, 10, txt=class_val, border=1)
                    pdf.cell(20, 10, txt=section, border=1)
                    pdf.cell(20, 10, txt=roll_no, border=1)
                    pdf.cell(30, 10, txt=voting_code, border=1, ln=True)
                
                conn.commit()
                
                pdf.ln(10)
                pdf.set_font("Arial", 'B', size=10)
                pdf.cell(200, 10, txt=f"Students Added: {students_added}", ln=True)
                pdf.cell(200, 10, txt=f"Students Updated: {students_updated}", ln=True)
                pdf.cell(200, 10, txt=f"Total Processed: {students_added + students_updated}", ln=True)
                pdf_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
                pdf_temp.close()
                
                pdf.output(pdf_temp.name)
                
                return send_file(
                    pdf_temp.name,
                    mimetype='application/pdf',
                    as_attachment=True,
                    download_name='student_voting_codes.pdf'
                )
        
        except Exception as inner_e:
            logging.error(f"Error processing Excel file: {str(inner_e)}")
            return jsonify({
                'success': False, 
                'message': f'Error processing Excel file: {str(inner_e)}'
            }), 500
            
    except Exception as e:
        logging.error(f"Error processing Excel file: {str(e)}")
        return jsonify({'success': False, 'message': f'Error processing Excel file: {str(e)}'}), 500
    
    finally:
        if temp_file and os.path.exists(temp_file.name):
            try:
                os.unlink(temp_file.name)
            except:
                pass

@app.route('/admin/student/reset/<int:student_id>', methods=['POST'])
def reset_student_code(student_id):
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        with DBContextManager() as conn:
            student = conn.execute('SELECT * FROM voters WHERE id = ? AND is_teacher = 0', (student_id,)).fetchone()
            
            if not student:
                logging.warning(f"Student ID {student_id} not found during reset attempt")
                return jsonify({'success': False, 'message': 'Student not found'})
            
            new_code = generate_unique_code()
            conn.execute('UPDATE voters SET voting_code = ? WHERE id = ?', (new_code, student_id))
            
            # Log the action
            logging.info(f"Admin {session.get('admin_username')} reset student code for student ID {student_id}")
            
            return jsonify({'success': True, 'new_code': new_code, 'message': 'Student code reset successfully'})
    except Exception as e:
        logging.error(f"Error in reset_student_code for ID {student_id}: {e}")
        return jsonify({'success': False, 'message': f'Error resetting student code: {str(e)}'})

@app.route('/admin/students/reset_all', methods=['POST'])
def reset_all_student_codes():
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        with DBContextManager() as conn:
            students = conn.execute('SELECT id FROM voters WHERE is_teacher = 0').fetchall()
            
            if not students:
                logging.warning("No students found to reset codes")
                return jsonify({'success': False, 'message': 'No students found'})
            
            updated_students = []
            for student in students:
                try:
                    new_code = generate_unique_code()
                    conn.execute('UPDATE voters SET voting_code = ? WHERE id = ?', (new_code, student['id']))
                    updated_students.append({'id': student['id'], 'new_code': new_code})
                except Exception as inner_e:
                    logging.error(f"Error updating student ID {student['id']}: {inner_e}")
            
            # Make sure to commit the changes
            conn.commit()
            
            # Log the action
            logging.info(f"Admin {session.get('admin_username')} reset all student codes - {len(updated_students)} codes updated")
            
            return jsonify({
                'success': True, 
                'updated_students': updated_students,
                'message': f'Reset {len(updated_students)} student codes successfully'
            })
    except Exception as e:
        logging.error(f"Error in reset_all_student_codes: {e}")
        return jsonify({'success': False, 'message': f'Error resetting student codes: {str(e)}'})

# Teacher Management Routes
@app.route('/admin/teachers')
def admin_teachers():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    
    try:
        with DBContextManager() as conn:
            teachers = conn.execute('SELECT * FROM voters WHERE is_teacher = 1').fetchall()
            return render_template('admin/teachers.html', teachers=teachers)
    except Exception as e:
        flash(f'Error: {str(e)}')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/teacher/add', methods=['POST'])
def add_teacher():
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    name = request.form.get('name', '').strip()
    subject = request.form.get('subject', '').strip()
    
    if not name:
        return jsonify({'success': False, 'message': 'Teacher name is required'})
    
    if len(name) > 100:
        return jsonify({'success': False, 'message': 'Teacher name is too long (maximum 100 characters)'})
    
    if len(subject) > 100:
        return jsonify({'success': False, 'message': 'Subject is too long (maximum 100 characters)'})
    
    voting_code = generate_unique_code()
    
    try:
        with DBContextManager() as conn:
            cursor = conn.execute(
                'INSERT INTO voters (name, subject, voting_code, is_teacher, has_voted) VALUES (?, ?, ?, 1, 0)',
                (name, subject, voting_code)
            )
            
            teacher_id = cursor.lastrowid
            teacher = conn.execute('SELECT * FROM voters WHERE id = ?', (teacher_id,)).fetchone()
            
            teacher_dict = {
                'id': teacher['id'],
                'name': teacher['name'],
                'subject': teacher['subject'],
                'voting_code': teacher['voting_code'],
                'has_voted': teacher['has_voted']
            }
            
            return jsonify({
                'success': True, 
                'teacher': teacher_dict,
                'message': f'Teacher {name} added successfully'
            })
    except Exception as e:
        logging.error(f"Error in add_teacher: {e}")
        return jsonify({'success': False, 'message': f'Error adding teacher: {str(e)}'})

@app.route('/admin/teacher/delete/<int:teacher_id>', methods=['POST'])
def delete_teacher(teacher_id):
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        with DBContextManager() as conn:
            teacher = conn.execute('SELECT * FROM voters WHERE id = ? AND is_teacher = 1', (teacher_id,)).fetchone()
            
            if not teacher:
                return jsonify({'success': False, 'message': 'Teacher not found'})
            
            if teacher['has_voted']:
                return jsonify({'success': False, 'message': 'Cannot delete a teacher who has already voted'})
            
            conn.execute('DELETE FROM voters WHERE id = ?', (teacher_id,))
            conn.commit()
            
            return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/teacher/reset/<int:teacher_id>', methods=['POST'])
def reset_teacher_code(teacher_id):
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        with DBContextManager() as conn:
            teacher = conn.execute('SELECT * FROM voters WHERE id = ? AND is_teacher = 1', (teacher_id,)).fetchone()
            
            if not teacher:
                logging.warning(f"Teacher ID {teacher_id} not found during reset attempt")
                return jsonify({'success': False, 'message': 'Teacher not found'})
            
            new_code = generate_unique_code()
            conn.execute('UPDATE voters SET voting_code = ? WHERE id = ?', (new_code, teacher_id))
            
            # Log the action
            logging.info(f"Admin {session.get('admin_username')} reset teacher code for teacher ID {teacher_id}")
            
            return jsonify({'success': True, 'new_code': new_code, 'message': 'Teacher code reset successfully'})
    except Exception as e:
        logging.error(f"Error in reset_teacher_code for ID {teacher_id}: {e}")
        return jsonify({'success': False, 'message': f'Error resetting teacher code: {str(e)}'})

@app.route('/admin/teachers/reset_all', methods=['POST'])
def reset_all_teacher_codes():
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        with DBContextManager() as conn:
            teachers = conn.execute('SELECT id FROM voters WHERE is_teacher = 1').fetchall()
            
            if not teachers:
                logging.warning("No teachers found to reset codes")
                return jsonify({'success': False, 'message': 'No teachers found'})
            
            updated_teachers = []
            for teacher in teachers:
                try:
                    new_code = generate_unique_code()
                    conn.execute('UPDATE voters SET voting_code = ? WHERE id = ?', (new_code, teacher['id']))
                    updated_teachers.append({'id': teacher['id'], 'new_code': new_code})
                except Exception as inner_e:
                    logging.error(f"Error updating teacher ID {teacher['id']}: {inner_e}")
            
            # Make sure to commit the changes
            conn.commit()
            
            # Log the action
            logging.info(f"Admin {session.get('admin_username')} reset all teacher codes - {len(updated_teachers)} codes updated")
            
            return jsonify({
                'success': True, 
                'updated_teachers': updated_teachers,
                'message': f'Reset {len(updated_teachers)} teacher codes successfully'
            })
    except Exception as e:
        logging.error(f"Error in reset_all_teacher_codes: {e}")
        return jsonify({'success': False, 'message': f'Error resetting teacher codes: {str(e)}'})

@app.route('/admin/teachers/upload', methods=['POST'])
def upload_teachers():
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file part'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'}), 400
    
    if not file.filename.lower().endswith('.xlsx'):
        return jsonify({'success': False, 'message': 'File must be an Excel (.xlsx) file'}), 400
    
    # Check file size (limit to 5MB)
    file_size = file.content_length or 0
    if file_size > 5 * 1024 * 1024:  # 5MB
        return jsonify({'success': False, 'message': 'File size exceeds the 5MB limit'}), 400
    
    temp_file = None
    pdf_output = None
    
    try:
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        file.save(temp_file.name)
        temp_file.close()
        
        try:
            workbook = openpyxl.load_workbook(temp_file.name, read_only=True, data_only=True)
            sheet = workbook.active
            
            if sheet.max_row < 2:  
                os.unlink(temp_file.name) 
                return jsonify({'success': False, 'message': 'Excel file is empty or has no data rows'}), 400
            
            if sheet.max_row > 1000:  
                os.unlink(temp_file.name) 
                return jsonify({'success': False, 'message': 'Excel file has too many rows (limit: 1000)'}), 400
            
            header_row = [cell.value for cell in sheet[1]]
            
            required_columns = ['Name', 'Subject']
            for col in required_columns:
                if col not in header_row:
                    os.unlink(temp_file.name)  
                    return jsonify({
                        'success': False, 
                        'message': f'Missing required column: {col}. Please make sure your Excel file has the columns: {", ".join(required_columns)}'
                    }), 400
            
            name_col = header_row.index('Name') + 1
            subject_col = header_row.index('Subject') + 1
            
            teachers_added = 0
            teachers_updated = 0
            processed_teachers = set()
            
            with DBContextManager() as conn:
                pdf = FPDF()
                pdf.add_page()
                pdf.set_font("Arial", size=12)
                
                pdf.cell(200, 10, txt="Pathshala Election 2082 - Teacher Voting Codes", ln=True, align='C')
                pdf.set_font("Arial", 'B', size=14)
                pdf.cell(200, 10, txt="CONFIDENTIAL", ln=True, align='C')
                pdf.set_font("Arial", size=10)
                pdf.cell(200, 10, txt=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align='C')
                pdf.ln(10)
                
                pdf.set_font("Arial", 'B', size=10)
                pdf.cell(100, 10, txt="Name", border=1)
                pdf.cell(50, 10, txt="Subject", border=1)
                pdf.cell(30, 10, txt="Voting Code", border=1, ln=True)
                
                pdf.set_font("Arial", size=10)
                
                for row in range(2, sheet.max_row + 1):
                    name = sheet.cell(row=row, column=name_col).value
                    subject = sheet.cell(row=row, column=subject_col).value
                    
                    if not name:
                        continue
                    
                    name = str(name)[:100].strip() 
                    subject = str(subject)[:100].strip() if subject is not None else ''
                    
                    # Using name as unique key for teachers
                    teacher_key = name.lower()
                    if teacher_key in processed_teachers:
                        continue
                    
                    processed_teachers.add(teacher_key)
                    
                    voting_code = generate_unique_code()
                    
                    existing = conn.execute(
                        'SELECT id, has_voted FROM voters WHERE name = ? AND is_teacher = 1',
                        (name,)
                    ).fetchone()
                    
                    if existing:
                        # Only update if teacher hasn't voted
                        if not existing['has_voted']:
                            conn.execute(
                                'UPDATE voters SET voting_code = ?, subject = ? WHERE id = ?',
                                (voting_code, subject, existing['id'])
                            )
                            teachers_updated += 1
                        else:
                            # Get the existing voting code if teacher has voted
                            teacher_data = conn.execute(
                                'SELECT voting_code FROM voters WHERE id = ?',
                                (existing['id'],)
                            ).fetchone()
                            voting_code = teacher_data['voting_code']
                    else:
                        conn.execute(
                            'INSERT INTO voters (name, subject, voting_code, is_teacher, has_voted) VALUES (?, ?, ?, 1, 0)',
                            (name, subject, voting_code)
                        )
                        teachers_added += 1
                    
                    pdf.cell(100, 10, txt=name, border=1)
                    pdf.cell(50, 10, txt=subject, border=1)
                    pdf.cell(30, 10, txt=voting_code, border=1, ln=True)
                
                conn.commit()
                
                pdf.ln(10)
                pdf.set_font("Arial", 'B', size=10)
                pdf.cell(200, 10, txt=f"Teachers Added: {teachers_added}", ln=True)
                pdf.cell(200, 10, txt=f"Teachers Updated: {teachers_updated}", ln=True)
                pdf.cell(200, 10, txt=f"Total Processed: {teachers_added + teachers_updated}", ln=True)
                pdf_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
                pdf_temp.close()
                
                pdf.output(pdf_temp.name)
                
                return send_file(
                    pdf_temp.name,
                    mimetype='application/pdf',
                    as_attachment=True,
                    download_name='teacher_voting_codes.pdf'
                )
        
        except Exception as inner_e:
            logging.error(f"Error processing Excel file: {str(inner_e)}")
            return jsonify({
                'success': False, 
                'message': f'Error processing Excel file: {str(inner_e)}'
            }), 500
            
    except Exception as e:
        logging.error(f"Error processing Excel file: {str(e)}")
        return jsonify({'success': False, 'message': f'Error processing Excel file: {str(e)}'}), 500
    
    finally:
        if temp_file and os.path.exists(temp_file.name):
            try:
                os.unlink(temp_file.name)
            except:
                pass

# Candidate Management Routes
@app.route('/admin/candidates')
def admin_candidates():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    
    try:
        with DBContextManager() as conn:
            candidates = conn.execute('''
                SELECT c.id, c.name, c.gender, c.image_path, c.logo_path, c.slogan, c.description,
                       COUNT(v.id) as vote_count
                FROM candidates c
                LEFT JOIN votes v ON c.id = v.candidate_id
                GROUP BY c.id
                ORDER BY c.gender, c.name
            ''').fetchall()
            
            candidates_list = [dict(candidate) for candidate in candidates]
            
            return render_template('admin/candidates.html', candidates=candidates_list)
    except sqlite3.Error as e:
        logging.error(f"Database error in admin_candidates: {str(e)}")
        flash(f'Database error: {str(e)}')
        return redirect(url_for('admin_dashboard'))
    except Exception as e:
        logging.error(f"Error in admin_candidates: {str(e)}")
        flash(f'Error loading candidates: {str(e)}')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/candidate/add', methods=['POST'])
def add_candidate():
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        logging.debug(f"Add candidate form data: {request.form}")
        logging.debug(f"Add candidate files: {request.files.keys()}")
        
        name = request.form.get('name', '').strip()
        gender = request.form.get('gender', '').strip()
        slogan = request.form.get('slogan', '').strip()
        description = request.form.get('description', '').strip()
        
        logging.debug(f"Candidate data - name: '{name}', gender: '{gender}'")
        
        if not name or not gender:
            logging.warning(f"Validation failed: name or gender missing - name: '{name}', gender: '{gender}'")
            return jsonify({'success': False, 'message': 'Name and gender are required'}), 400
        
        if len(name) > 100:
            logging.warning(f"Validation failed: name too long - length: {len(name)}")
            return jsonify({'success': False, 'message': 'Candidate name is too long (maximum 100 characters)'}), 400
        
        if gender not in ['Male', 'Female']:
            logging.warning(f"Validation failed: invalid gender - got: '{gender}'")
            return jsonify({'success': False, 'message': 'Gender must be either Male or Female'}), 400
        
        if len(slogan) > 200:
            logging.warning(f"Validation failed: slogan too long - length: {len(slogan)}")
            return jsonify({'success': False, 'message': 'Slogan is too long (maximum 200 characters)'}), 400
        
        if len(description) > 1000:
            logging.warning(f"Validation failed: description too long - length: {len(description)}")
            return jsonify({'success': False, 'message': 'Description is too long (maximum 1000 characters)'}), 400
        
        image_path = None
        logo_path = None
        saved_files = []
        
        if 'image' in request.files and request.files['image'].filename:
            image = request.files['image']
            logging.debug(f"Processing image: {image.filename}")
            
            allowed_extensions = {'jpg', 'jpeg', 'png', 'gif'}
            file_ext = ''
            if '.' in image.filename:
                file_ext = image.filename.rsplit('.', 1)[1].lower()
                
            logging.debug(f"Image extension: {file_ext}")
            
            if not '.' in image.filename or file_ext not in allowed_extensions:
                logging.warning(f"Invalid image format: {image.filename}")
                return jsonify({'success': False, 'message': f'Invalid image format: {file_ext}. Allowed formats: jpg, jpeg, png, gif'}), 400
            
            try:
                timestamp = int(datetime.now().timestamp())
                image_filename = secure_filename(f"{name.replace(' ', '_')}_{timestamp}{os.path.splitext(image.filename)[1]}")
                image_path = os.path.join('static', 'img', 'candidates', image_filename)
                
                os.makedirs(os.path.dirname(image_path), exist_ok=True)
                
                image.save(image_path)
                saved_files.append(image_path)
                logging.debug(f"Image saved to: {image_path}")
            except Exception as img_error:
                logging.error(f"Error saving image: {str(img_error)}")
                return jsonify({'success': False, 'message': f'Error saving image: {str(img_error)}'}), 500
        
        if 'logo' in request.files and request.files['logo'].filename:
            logo = request.files['logo']
            logging.debug(f"Processing logo: {logo.filename}")
            
            allowed_extensions = {'jpg', 'jpeg', 'png', 'gif'}
            file_ext = ''
            if '.' in logo.filename:
                file_ext = logo.filename.rsplit('.', 1)[1].lower()
                
            logging.debug(f"Logo extension: {file_ext}")
            
            if not '.' in logo.filename or file_ext not in allowed_extensions:
                for file_path in saved_files:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                logging.warning(f"Invalid logo format: {logo.filename}")
                return jsonify({'success': False, 'message': f'Invalid logo format: {file_ext}. Allowed formats: jpg, jpeg, png, gif'}), 400
            
            try:
                timestamp = int(datetime.now().timestamp())
                logo_filename = secure_filename(f"{name.replace(' ', '_')}_logo_{timestamp}{os.path.splitext(logo.filename)[1]}")
                logo_path = os.path.join('static', 'img', 'logos', logo_filename)
                
                os.makedirs(os.path.dirname(logo_path), exist_ok=True)
                
                logo.save(logo_path)
                saved_files.append(logo_path)
                logging.debug(f"Logo saved to: {logo_path}")
            except Exception as logo_error:
                for file_path in saved_files:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                logging.error(f"Error saving logo: {str(logo_error)}")
                return jsonify({'success': False, 'message': f'Error saving logo: {str(logo_error)}'}), 500
        
        logging.debug("Proceeding with database operations")
        with DBContextManager() as conn:
            conn.execute('BEGIN TRANSACTION')
            
            try:
                cursor = conn.execute(
                    'INSERT INTO candidates (name, gender, image_path, logo_path, slogan, description) VALUES (?, ?, ?, ?, ?, ?)',
                    (name, gender, image_path, logo_path, slogan, description)
                )
                
                candidate_id = cursor.lastrowid
                logging.debug(f"Inserted candidate with ID: {candidate_id}")
                
                candidate = conn.execute('SELECT * FROM candidates WHERE id = ?', (candidate_id,)).fetchone()
                
                if not candidate:
                    conn.rollback()
                    for file_path in saved_files:
                        if os.path.exists(file_path):
                            os.remove(file_path)
                    logging.error("Failed to retrieve newly inserted candidate")
                    return jsonify({'success': False, 'message': 'Failed to create candidate record'}), 500
                
                conn.commit()
                logging.info(f"Successfully added candidate {name} (ID: {candidate_id})")
                
                return jsonify({
                    'success': True, 
                    'candidate': dict(candidate),
                    'message': f'Candidate {name} added successfully'
                })
            except Exception as db_error:
                conn.rollback()
                for file_path in saved_files:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                logging.error(f"Database error in candidate insert: {str(db_error)}")
                return jsonify({'success': False, 'message': f'Database error: {str(db_error)}'}), 500
            
    except sqlite3.Error as e:
        logging.error(f"Database error in add_candidate: {str(e)}")
        if 'saved_files' in locals():
            for file_path in saved_files:
                if os.path.exists(file_path):
                    os.remove(file_path)
        return jsonify({'success': False, 'message': f'Database error: {str(e)}'}), 500
    except Exception as e:
        logging.error(f"Error in add_candidate: {str(e)}")
        if 'saved_files' in locals():
            for file_path in saved_files:
                if os.path.exists(file_path):
                    os.remove(file_path)
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/admin/candidate/edit/<int:candidate_id>', methods=['POST'])
def edit_candidate(candidate_id):
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        with DBContextManager() as conn:
            candidate = conn.execute('SELECT * FROM candidates WHERE id = ?', (candidate_id,)).fetchone()
            
            if not candidate:
                return jsonify({'success': False, 'message': 'Candidate not found'})
            
            name = request.form.get('name', candidate['name'])
            slogan = request.form.get('slogan', candidate['slogan'])
            description = request.form.get('description', candidate['description'])
            
            image_path = candidate['image_path']
            if 'image' in request.files and request.files['image'].filename:
                image = request.files['image']
                if image.filename:
                    if image_path and os.path.exists(image_path):
                        os.remove(image_path)
                    
                    image_filename = secure_filename(f"{name.replace(' ', '_')}_{int(datetime.now().timestamp())}{os.path.splitext(image.filename)[1]}")
                    image_path = os.path.join('static', 'img', 'candidates', image_filename)
                    os.makedirs(os.path.dirname(image_path), exist_ok=True)
                    image.save(image_path)
            
            logo_path = candidate['logo_path']
            if 'logo' in request.files and request.files['logo'].filename:
                logo = request.files['logo']
                if logo.filename:
                    if logo_path and os.path.exists(logo_path):
                        os.remove(logo_path)
                    
                    logo_filename = secure_filename(f"{name.replace(' ', '_')}_logo_{int(datetime.now().timestamp())}{os.path.splitext(logo.filename)[1]}")
                    logo_path = os.path.join('static', 'img', 'logos', logo_filename)
                    os.makedirs(os.path.dirname(logo_path), exist_ok=True)
                    logo.save(logo_path)
            
            conn.execute(
                'UPDATE candidates SET name = ?, image_path = ?, logo_path = ?, slogan = ?, description = ? WHERE id = ?',
                (name, image_path, logo_path, slogan, description, candidate_id)
            )
            conn.commit()
            
            updated_candidate = conn.execute('SELECT * FROM candidates WHERE id = ?', (candidate_id,)).fetchone()
            
            return jsonify({
                'success': True,
                'candidate': dict(updated_candidate)
            })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/candidate/delete/<int:candidate_id>', methods=['POST'])
def delete_candidate(candidate_id):
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        with DBContextManager() as conn:
            candidate = conn.execute('SELECT * FROM candidates WHERE id = ?', (candidate_id,)).fetchone()
            
            if not candidate:
                return jsonify({'success': False, 'message': 'Candidate not found'})

            votes = conn.execute('SELECT COUNT(*) FROM votes WHERE candidate_id = ?', (candidate_id,)).fetchone()[0]
            if votes > 0:
                return jsonify({'success': False, 'message': 'Cannot delete a candidate who has received votes'})

            if candidate['image_path'] and os.path.exists(candidate['image_path']):
                os.remove(candidate['image_path'])
            
            if candidate['logo_path'] and os.path.exists(candidate['logo_path']):
                os.remove(candidate['logo_path'])
            
            conn.execute('DELETE FROM candidates WHERE id = ?', (candidate_id,))
            conn.commit()
            
            return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

# Results page
@app.route('/admin/results')
def admin_results():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    
    try:
        with DBContextManager() as conn:
            view_exists = conn.execute("SELECT name FROM sqlite_master WHERE type='view' AND name='vote_weights'").fetchone()
            
            if not view_exists:
                conn.execute('''
                    CREATE VIEW vote_weights AS
                    SELECT
                        v.id,
                        v.voter_id,
                        v.candidate_id,
                        v.timestamp,
                        CASE WHEN vr.is_teacher = 1 THEN 6 ELSE 1 END as vote_weight
                    FROM votes v
                    JOIN voters vr ON v.voter_id = vr.id
                ''')
            
            male_candidates = conn.execute('''
                SELECT c.id, c.name, c.image_path, c.logo_path, c.gender,
                       COUNT(v.id) as votes,
                       SUM(CASE WHEN vr.is_teacher = 1 THEN 6 ELSE 1 END) as points
                FROM candidates c
                LEFT JOIN votes v ON c.id = v.candidate_id
                LEFT JOIN voters vr ON v.voter_id = vr.id
                WHERE c.gender = 'Male'
                GROUP BY c.id
                ORDER BY points DESC, votes DESC, c.name
            ''').fetchall()

            female_candidates = conn.execute('''
                SELECT c.id, c.name, c.image_path, c.logo_path, c.gender,
                       COUNT(v.id) as votes,
                       SUM(CASE WHEN vr.is_teacher = 1 THEN 6 ELSE 1 END) as points
                FROM candidates c
                LEFT JOIN votes v ON c.id = v.candidate_id
                LEFT JOIN voters vr ON v.voter_id = vr.id
                WHERE c.gender = 'Female'
                GROUP BY c.id
                ORDER BY points DESC, votes DESC, c.name
            ''').fetchall()
            
            male_results = [dict(candidate) for candidate in male_candidates]
            female_results = [dict(candidate) for candidate in female_candidates]
            
            total_students = conn.execute('SELECT COUNT(*) FROM voters WHERE is_teacher = 0').fetchone()[0]
            total_teachers = conn.execute('SELECT COUNT(*) FROM voters WHERE is_teacher = 1').fetchone()[0]
            
            students_voted = conn.execute('SELECT COUNT(*) FROM voters WHERE is_teacher = 0 AND has_voted = 1').fetchone()[0]
            teachers_voted = conn.execute('SELECT COUNT(*) FROM voters WHERE is_teacher = 1 AND has_voted = 1').fetchone()[0]
            
            student_vote_percent = (students_voted / total_students * 100) if total_students > 0 else 0
            teacher_vote_percent = (teachers_voted / total_teachers * 100) if total_teachers > 0 else 0
            
            total_voters = total_students + total_teachers
            total_voted = students_voted + teachers_voted
            percentage_voted = (total_voted / total_voters * 100) if total_voters > 0 else 0
            remaining_voters = total_voters - total_voted
            
            total_points = (teachers_voted * 6) + students_voted
            
            class_data = conn.execute('''
                SELECT class, section, 
                       COUNT(*) as total_students, 
                       SUM(has_voted) as voted_students
                FROM voters 
                WHERE is_teacher = 0 
                GROUP BY class, section
                ORDER BY class, section
            ''').fetchall()
            
            class_labels = [f"Class {c['class']}{c['section']}" for c in class_data]
            class_percentages = [(c['voted_students'] / c['total_students'] * 100) if c['total_students'] > 0 else 0 for c in class_data]
            
            male_winner = male_results[0] if male_results and male_results[0]['votes'] > 0 else None
            female_winner = female_results[0] if female_results and female_results[0]['votes'] > 0 else None
            
            return render_template('admin/results.html',
                                male_results=male_results,
                                female_results=female_results,
                                male_winner=male_winner,
                                female_winner=female_winner,
                                total_students=total_students,
                                total_teachers=total_teachers,
                                students_voted=students_voted,
                                teachers_voted=teachers_voted,
                                student_vote_percent=student_vote_percent,
                                teacher_vote_percent=teacher_vote_percent,
                                percentage_voted=percentage_voted,
                                remaining_voters=remaining_voters,
                                total_votes=total_voted,
                                total_points=total_points,
                                class_labels=class_labels,
                                class_percentages=class_percentages)
    except Exception as e:
        flash(f'Error: {str(e)}')
        return redirect(url_for('admin_dashboard'))

@app.route('/admin/export_results')
def export_results():
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    try:
        with DBContextManager() as conn:
            male_candidates = conn.execute('''
                SELECT c.id, c.name, c.gender, COUNT(v.id) as votes
                FROM candidates c
                LEFT JOIN votes v ON c.id = v.candidate_id
                WHERE c.gender = 'Male'
                GROUP BY c.id
                ORDER BY votes DESC
            ''').fetchall()
            
            female_candidates = conn.execute('''
                SELECT c.id, c.name, c.gender, COUNT(v.id) as votes
                FROM candidates c
                LEFT JOIN votes v ON c.id = v.candidate_id
                WHERE c.gender = 'Female'
                GROUP BY c.id
                ORDER BY votes DESC
            ''').fetchall()
            
            total_students = conn.execute('SELECT COUNT(*) FROM voters WHERE is_teacher = 0').fetchone()[0]
            total_teachers = conn.execute('SELECT COUNT(*) FROM voters WHERE is_teacher = 1').fetchone()[0]
            
            students_voted = conn.execute('SELECT COUNT(*) FROM voters WHERE is_teacher = 0 AND has_voted = 1').fetchone()[0]
            teachers_voted = conn.execute('SELECT COUNT(*) FROM voters WHERE is_teacher = 1 AND has_voted = 1').fetchone()[0]
            
            class_data = conn.execute('''
                SELECT class, section, 
                       COUNT(*) as total_students, 
                       SUM(has_voted) as voted_students
                FROM voters 
                WHERE is_teacher = 0 
                GROUP BY class, section
                ORDER BY class, section
            ''').fetchall()
            
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            
            pdf.cell(200, 10, txt="Pathshala Election 2082 - Election Results", ln=True, align='C')
            pdf.set_font("Arial", 'B', size=14)
            pdf.cell(200, 10, txt="RESULTS SUMMARY", ln=True, align='C')
            pdf.set_font("Arial", size=10)
            pdf.cell(200, 10, txt=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align='C')
            pdf.ln(10)
            
            pdf.set_font("Arial", 'B', size=12)
            pdf.cell(200, 10, txt="Voting Statistics", ln=True)
            pdf.ln(5)
            
            pdf.set_font("Arial", size=10)
            
            pdf.cell(100, 8, txt="Total Students", border=1)
            pdf.cell(30, 8, txt=str(total_students), border=1, ln=True)
            
            pdf.cell(100, 8, txt="Students who voted", border=1)
            pdf.cell(30, 8, txt=str(students_voted), border=1, ln=True)
            
            pdf.cell(100, 8, txt="Student voting percentage", border=1)
            student_percent = (students_voted / total_students * 100) if total_students > 0 else 0
            pdf.cell(30, 8, txt=f"{student_percent:.1f}%", border=1, ln=True)
            
            pdf.cell(100, 8, txt="Total Teachers", border=1)
            pdf.cell(30, 8, txt=str(total_teachers), border=1, ln=True)
            
            pdf.cell(100, 8, txt="Teachers who voted", border=1)
            pdf.cell(30, 8, txt=str(teachers_voted), border=1, ln=True)
            
            pdf.cell(100, 8, txt="Teacher voting percentage", border=1)
            teacher_percent = (teachers_voted / total_teachers * 100) if total_teachers > 0 else 0
            pdf.cell(30, 8, txt=f"{teacher_percent:.1f}%", border=1, ln=True)
            
            pdf.cell(100, 8, txt="Total participation", border=1)
            total_percent = ((students_voted + teachers_voted) / (total_students + total_teachers) * 100) if (total_students + total_teachers) > 0 else 0
            pdf.cell(30, 8, txt=f"{total_percent:.1f}%", border=1, ln=True)
            
            pdf.ln(10)
            
            pdf.set_font("Arial", 'B', size=12)
            pdf.cell(200, 10, txt="Male Candidates Results", ln=True)
            pdf.ln(5)
            
            pdf.set_font("Arial", 'B', size=10)
            pdf.cell(10, 8, txt="Rank", border=1)
            pdf.cell(100, 8, txt="Candidate Name", border=1)
            pdf.cell(30, 8, txt="Votes", border=1)
            pdf.cell(30, 8, txt="Percentage", border=1, ln=True)
            
            pdf.set_font("Arial", size=10)
            total_male_votes = sum(c['votes'] for c in male_candidates)
            
            for i, candidate in enumerate(male_candidates):
                rank = i + 1
                name = candidate['name']
                votes = candidate['votes']
                percentage = (votes / total_male_votes * 100) if total_male_votes > 0 else 0
                
                pdf.cell(10, 8, txt=str(rank), border=1)
                pdf.cell(100, 8, txt=name, border=1)
                pdf.cell(30, 8, txt=str(votes), border=1)
                pdf.cell(30, 8, txt=f"{percentage:.1f}%", border=1, ln=True)
            
            pdf.ln(10)
            
            pdf.set_font("Arial", 'B', size=12)
            pdf.cell(200, 10, txt="Female Candidates Results", ln=True)
            pdf.ln(5)
            
            pdf.set_font("Arial", 'B', size=10)
            pdf.cell(10, 8, txt="Rank", border=1)
            pdf.cell(100, 8, txt="Candidate Name", border=1)
            pdf.cell(30, 8, txt="Votes", border=1)
            pdf.cell(30, 8, txt="Percentage", border=1, ln=True)
            
            pdf.set_font("Arial", size=10)
            total_female_votes = sum(c['votes'] for c in female_candidates)
            
            for i, candidate in enumerate(female_candidates):
                rank = i + 1
                name = candidate['name']
                votes = candidate['votes']
                percentage = (votes / total_female_votes * 100) if total_female_votes > 0 else 0
                
                pdf.cell(10, 8, txt=str(rank), border=1)
                pdf.cell(100, 8, txt=name, border=1)
                pdf.cell(30, 8, txt=str(votes), border=1)
                pdf.cell(30, 8, txt=f"{percentage:.1f}%", border=1, ln=True)
            
            pdf.ln(10)
            
            pdf.set_font("Arial", 'B', size=12)
            pdf.cell(200, 10, txt="Class-wise Voting Statistics", ln=True)
            pdf.ln(5)
            
            pdf.set_font("Arial", 'B', size=10)
            pdf.cell(40, 8, txt="Class", border=1)
            pdf.cell(40, 8, txt="Total Students", border=1)
            pdf.cell(40, 8, txt="Voted", border=1)
            pdf.cell(40, 8, txt="Percentage", border=1, ln=True)
            
            pdf.set_font("Arial", size=10)
            for data in class_data:
                class_name = f"Class {data['class']}{data['section']}"
                total = data['total_students']
                voted = data['voted_students']
                percentage = (voted / total * 100) if total > 0 else 0
                
                pdf.cell(40, 8, txt=class_name, border=1)
                pdf.cell(40, 8, txt=str(total), border=1)
                pdf.cell(40, 8, txt=str(voted), border=1)
                pdf.cell(40, 8, txt=f"{percentage:.1f}%", border=1, ln=True)
            
            pdf_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
            pdf_temp.close()
            
            pdf.output(pdf_temp.name)
            
            return send_file(
                pdf_temp.name,
                mimetype='application/pdf',
                as_attachment=True,
                download_name='election_results.pdf'
            )
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error generating results: {str(e)}'}), 500

@app.route('/admin/refresh_results')
def refresh_results():
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    return jsonify({'success': True})

# Danger Zone - Database Reset
@app.route('/admin/reset')
def admin_reset():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    
    return render_template('admin/reset.html')

@app.route('/admin/reset/confirm', methods=['POST'])
def admin_reset_confirm():
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    password = request.form.get('password')
    
    with DBContextManager() as conn:
        admin = conn.execute('SELECT password_hash FROM admin WHERE id = ?', (session['admin_id'],)).fetchone()
        
        if not admin or not check_password_hash(admin['password_hash'], password):
            return jsonify({'success': False, 'message': 'Incorrect admin password'})
    
    try:
        init_db()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@app.route('/admin/students/download')
def download_all_students():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    
    try:
        with DBContextManager() as conn:
            students = conn.execute('SELECT * FROM voters WHERE is_teacher = 0 ORDER BY class, section, roll_no').fetchall()
            
            pdf = FPDF()
            pdf.add_page(orientation='L')  
            pdf.set_font("Arial", size=12)
            
            pdf.cell(280, 10, txt="Pathshala Election 2082 - All Student Voting Codes", ln=True, align='C')
            pdf.set_font("Arial", 'B', size=14)
            pdf.cell(280, 10, txt="CONFIDENTIAL", ln=True, align='C')
            pdf.set_font("Arial", size=10)
            pdf.cell(280, 10, txt=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align='C')
            pdf.ln(10)
            
            pdf.set_font("Arial", 'B', size=10)
            pdf.cell(10, 10, txt="ID", border=1)
            pdf.cell(60, 10, txt="Name", border=1)
            pdf.cell(20, 10, txt="Class", border=1)
            pdf.cell(20, 10, txt="Section", border=1) 
            pdf.cell(20, 10, txt="Roll No", border=1)
            pdf.cell(30, 10, txt="Voting Code", border=1)
            pdf.cell(30, 10, txt="Vote Status", border=1, ln=True)
            
            pdf.set_font("Arial", size=10)
            students_per_page = 25
            student_count = 0
            
            for student in students:
                pdf.cell(10, 10, txt=str(student['id']), border=1)
                pdf.cell(60, 10, txt=student['name'], border=1)
                pdf.cell(20, 10, txt=str(student['class'] or ''), border=1)
                pdf.cell(20, 10, txt=str(student['section'] or ''), border=1)
                pdf.cell(20, 10, txt=str(student['roll_no'] or ''), border=1)
                pdf.cell(30, 10, txt=student['voting_code'], border=1)
                vote_status = "Voted" if student['has_voted'] else "Not Voted"
                pdf.cell(30, 10, txt=vote_status, border=1, ln=True)
                
                student_count += 1
                if student_count % students_per_page == 0 and student_count < len(students):
                    pdf.add_page(orientation='L')
                    
                    pdf.set_font("Arial", 'B', size=10)
                    pdf.cell(10, 10, txt="ID", border=1)
                    pdf.cell(60, 10, txt="Name", border=1)
                    pdf.cell(20, 10, txt="Class", border=1)
                    pdf.cell(20, 10, txt="Section", border=1)
                    pdf.cell(20, 10, txt="Roll No", border=1)
                    pdf.cell(30, 10, txt="Voting Code", border=1)
                    pdf.cell(30, 10, txt="Vote Status", border=1, ln=True)
                    
                    pdf.set_font("Arial", size=10)
            
            pdf.add_page()
            pdf.set_font("Arial", 'B', size=12)
            pdf.cell(280, 10, txt="Class-wise Voting Statistics", ln=True, align='C')
            pdf.ln(5)
            
            class_stats = conn.execute('''
                SELECT class, section, 
                       COUNT(*) as total, 
                       SUM(has_voted) as voted
                FROM voters 
                WHERE is_teacher = 0 
                GROUP BY class, section
                ORDER BY class, section
            ''').fetchall()
            
            pdf.set_font("Arial", 'B', size=10)
            pdf.cell(40, 10, txt="Class", border=1)
            pdf.cell(40, 10, txt="Section", border=1)
            pdf.cell(40, 10, txt="Total Students", border=1)
            pdf.cell(40, 10, txt="Voted", border=1)
            pdf.cell(40, 10, txt="Percentage", border=1, ln=True)
            
            pdf.set_font("Arial", size=10)
            for stat in class_stats:
                percent = (stat['voted'] / stat['total'] * 100) if stat['total'] > 0 else 0
                pdf.cell(40, 10, txt=str(stat['class'] or ''), border=1)
                pdf.cell(40, 10, txt=str(stat['section'] or ''), border=1)
                pdf.cell(40, 10, txt=str(stat['total']), border=1)
                pdf.cell(40, 10, txt=str(stat['voted']), border=1)
                pdf.cell(40, 10, txt=f"{percent:.1f}%", border=1, ln=True)
            
            pdf_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
            pdf_temp.close()
            
            pdf.output(pdf_temp.name)

            return send_file(
                pdf_temp.name,
                mimetype='application/pdf',
                as_attachment=True,
                download_name='all_student_voting_codes.pdf'
            )
            
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}')
        return redirect(url_for('admin_students'))

@app.route('/admin/voter/search')
def search_voter():
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    query = request.args.get('query', '').strip()
    
    if not query:
        return jsonify({'success': False, 'message': 'No search query provided'})
    
    conn = get_db()
    
    if query.isdigit():
        voter = conn.execute(
            'SELECT * FROM voters WHERE id = ?',
            (query,)
        ).fetchone()
        
        if voter:
            conn.close()
            return jsonify({
                'success': True, 
                'voter': dict(voter)
            })

    voter = conn.execute(
        'SELECT * FROM voters WHERE voting_code = ?',
        (query,)
    ).fetchone()
    
    if voter is None:
        voter = conn.execute(
            'SELECT * FROM voters WHERE voting_code = ?',
            (query.upper(),)
        ).fetchone()

    if voter is None:
        voter = conn.execute(
            'SELECT * FROM voters WHERE UPPER(voting_code) = UPPER(?)',
            (query,)
        ).fetchone()
    
    if voter:
        conn.close()
        return jsonify({
            'success': True, 
            'voter': dict(voter)
        })
    
    voters = conn.execute(
        'SELECT * FROM voters WHERE name LIKE ? ORDER BY is_teacher, class, section, name',
        (f'%{query}%',)
    ).fetchall()
    
    conn.close()
    
    if not voters:
        return jsonify({
            'success': False, 
            'message': 'No voters found matching your search'
        })
    
    if len(voters) == 1:
        return jsonify({
            'success': True, 
            'voter': dict(voters[0])
        })
    
    voters_list = [dict(v) for v in voters]
    return jsonify({
        'success': True, 
        'voters': voters_list
    })

@app.route('/admin/voter/search_advanced')
def search_voter_advanced():
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    query = request.args.get('query', '').strip()
    class_val = request.args.get('class', '').strip()
    section_val = request.args.get('section', '').strip()
    status_val = request.args.get('status', '').strip()
    
    where_clauses = []
    params = []
    
    if query:
        where_clauses.append("(name LIKE ? OR id = ? OR UPPER(voting_code) = UPPER(?))")
        params.extend([f'%{query}%', query, query])
    
    if class_val:
        where_clauses.append("class = ?")
        params.append(class_val)
    
    if section_val:
        where_clauses.append("section = ?")
        params.append(section_val)
    
    if status_val:
        if status_val == 'voted':
            where_clauses.append("has_voted = 1")
        elif status_val == 'not_voted':
            where_clauses.append("has_voted = 0")
    
    sql = "SELECT * FROM voters"
    if where_clauses:
        sql += " WHERE " + " AND ".join(where_clauses)
    
    sql += " ORDER BY is_teacher, class, section, name"
    
    conn = get_db()
    voters = conn.execute(sql, params).fetchall()
    conn.close()
    
    if not voters:
        return jsonify({
            'success': False, 
            'message': 'No voters found matching your search criteria'
        })
    
    voters_list = [dict(v) for v in voters]
    return jsonify({
        'success': True, 
        'voters': voters_list
    })

@app.route('/admin/revoke_vote', methods=['POST'])
def revoke_vote():
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    voter_id_or_code = request.form.get('voter_id')
    
    if not voter_id_or_code:
        return jsonify({'success': False, 'message': 'No voter ID or code provided'})
    
    try:
        with DBContextManager() as conn:
            if voter_id_or_code.isdigit():
                voter = conn.execute('SELECT * FROM voters WHERE id = ?', (voter_id_or_code,)).fetchone()
            else:
                voter = conn.execute('SELECT * FROM voters WHERE voting_code = ? OR UPPER(voting_code) = UPPER(?)', 
                                   (voter_id_or_code, voter_id_or_code)).fetchone()
            
            if not voter:
                logging.warning(f"Revoke vote attempted for non-existent voter: {voter_id_or_code}")
                return jsonify({'success': False, 'message': 'Voter not found'})
            
            if not voter['has_voted']:
                return jsonify({'success': False, 'message': 'This voter has not cast a vote yet'})
            conn.execute('BEGIN TRANSACTION')
            deleted = conn.execute('DELETE FROM votes WHERE voter_id = ?', (voter['id'],)).rowcount
            conn.execute('UPDATE voters SET has_voted = 0 WHERE id = ?', (voter['id'],))
            
            conn.commit()
            
            logging.info(f"Vote revoked for voter {voter['id']} (deleted {deleted} vote records)")
            
            return jsonify({
                'success': True, 
                'message': 'Vote revoked successfully',
                'voter': {
                    'id': voter['id'],
                    'name': voter['name'],
                    'class': voter['class'],
                    'section': voter['section'],
                    'roll_no': voter['roll_no'],
                    'voting_code': voter['voting_code']
                }
            })
    except sqlite3.Error as e:
        logging.error(f"Database error in revoke_vote: {str(e)}")
        return jsonify({'success': False, 'message': f'Database error: {str(e)}'})
    except Exception as e:
        logging.error(f"Error in revoke_vote: {str(e)}")
        return jsonify({'success': False, 'message': f'Error revoking vote: {str(e)}'})

@app.route('/admin/revoke_group_votes', methods=['POST'])
def revoke_group_votes():
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    class_val = request.form.get('class')
    section_val = request.form.get('section')
    password = request.form.get('password')
    
    if not password:
        return jsonify({'success': False, 'message': 'Password is required'})
    
    if not class_val and not section_val:
        return jsonify({'success': False, 'message': 'No class or section specified. Please select at least one filter.'})
    
    try:
        with DBContextManager() as conn:
            conn.execute('BEGIN EXCLUSIVE TRANSACTION')
            admin = conn.execute('SELECT password_hash FROM admin WHERE id = ?', (session['admin_id'],)).fetchone()
            
            if not admin:
                conn.rollback()
                return jsonify({'success': False, 'message': 'Admin account not found'}), 500
            
            if not check_password_hash(admin['password_hash'], password):
                conn.rollback()
                return jsonify({'success': False, 'message': 'Incorrect admin password'})
                
            where_clause = 'WHERE has_voted = 1'
            params = []
            
            if class_val:
                where_clause += ' AND class = ?'
                params.append(class_val)
            
            if section_val:
                where_clause += ' AND section = ?'
                params.append(section_val)
            
            voters_to_revoke = conn.execute(f'SELECT id FROM voters {where_clause}', params).fetchall()
            
            if not voters_to_revoke:
                conn.rollback()
                return jsonify({'success': False, 'message': 'No voters found matching the criteria'})
            
            voter_ids = [v['id'] for v in voters_to_revoke]
            placeholders = ','.join(['?'] * len(voter_ids))
            
            deleted_votes = conn.execute(
                f'DELETE FROM votes WHERE voter_id IN ({placeholders})',
                voter_ids
            ).rowcount
            
            updated_voters = conn.execute(
                f'UPDATE voters SET has_voted = 0 WHERE id IN ({placeholders})',
                voter_ids
            ).rowcount
            
            conn.commit()
            
            logging.info(f"Group vote revocation: {len(voter_ids)} voters, {deleted_votes} votes deleted by admin {session.get('admin_username')}")
            
            return jsonify({
                'success': True, 
                'count': len(voter_ids), 
                'votes_deleted': deleted_votes,
                'message': f'Successfully revoked {len(voter_ids)} votes'
            })
    except sqlite3.Error as e:
        logging.error(f"Database error in revoke_group_votes: {str(e)}")
        return jsonify({'success': False, 'message': f'Database error: {str(e)}'}), 500
    except Exception as e:
        logging.error(f"Error in revoke_group_votes: {str(e)}")
        return jsonify({'success': False, 'message': f'Error revoking votes: {str(e)}'}), 500

@app.route('/admin/reset_votes', methods=['POST'])
def reset_votes():
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    password = request.form.get('password')
    confirm_text = request.form.get('confirm_text')
    
    if not password:
        return jsonify({'success': False, 'message': 'Admin password is required'})
    
    if confirm_text != 'RESET VOTES':
        return jsonify({'success': False, 'message': 'Confirmation text does not match'})
    
    try:
        with DBContextManager() as conn:
            conn.execute('BEGIN EXCLUSIVE TRANSACTION')
            
            admin = conn.execute('SELECT password_hash FROM admin WHERE id = ?', (session['admin_id'],)).fetchone()
            if not admin:
                conn.rollback()
                return jsonify({'success': False, 'message': 'Admin account not found'}), 500
            
            if not check_password_hash(admin['password_hash'], password):
                conn.rollback()
                return jsonify({'success': False, 'message': 'Incorrect admin password'})
            
            votes_count = conn.execute('SELECT COUNT(*) as count FROM votes').fetchone()['count']
            
            conn.execute('DELETE FROM votes')
            conn.execute('UPDATE voters SET has_voted = 0')
            
            conn.commit()
            
            logging.info(f"Admin {session['admin_username']} reset {votes_count} votes")
            
            return jsonify({
                'success': True, 
                'message': f'All votes have been reset successfully. {votes_count} votes were deleted.'
            })
    except sqlite3.Error as e:
        logging.error(f"Database error in reset_votes: {str(e)}")
        return jsonify({
            'success': False, 
            'message': f'Database error: {str(e)}'
        }), 500
    except Exception as e:
        logging.error(f"Error in reset_votes: {str(e)}")
        return jsonify({
            'success': False, 
            'message': f'Error resetting votes: {str(e)}'
        }), 500

@app.route('/admin/factory_reset', methods=['POST'])
def factory_reset():
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    password = request.form.get('password')
    confirm_text = request.form.get('confirm_text')
    
    if confirm_text != 'FACTORY RESET':
        return jsonify({'success': False, 'message': 'Confirmation text does not match'})
    
    try:
        with DBContextManager() as conn:
            admin = conn.execute('SELECT password_hash FROM admin WHERE id = ?', (session['admin_id'],)).fetchone()
            
            if not admin or not check_password_hash(admin['password_hash'], password):
                return jsonify({'success': False, 'message': 'Incorrect admin password'})
            
            tables = conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
            table_names = [table['name'] for table in tables]
            
            if 'votes' in table_names:
                conn.execute('DELETE FROM votes')
                
            if 'candidates' in table_names:
                conn.execute('DELETE FROM candidates')
                
            if 'voters' in table_names:
                conn.execute('DELETE FROM voters')
                
            if 'election_settings' in table_names:
                conn.execute('DELETE FROM election_settings')
            
            for table in ['votes', 'candidates', 'voters', 'election_settings']:
                if table in table_names:
                    conn.execute(f"DELETE FROM sqlite_sequence WHERE name = '{table}'")
            
            conn.commit()
            
            return jsonify({
                'success': True, 
                'message': 'Factory reset completed successfully'
            })
    except Exception as e:
        return jsonify({
            'success': False, 
            'message': f'Error during factory reset: {str(e)}'
        })

@app.route('/admin/change_password', methods=['POST'])
def admin_change_password():
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')
    
    if not current_password or not new_password or not confirm_password:
        return jsonify({'success': False, 'message': 'All fields are required'})
    
    if new_password != confirm_password:
        return jsonify({'success': False, 'message': 'New passwords do not match'})
    
    if len(new_password) < 8:
        return jsonify({'success': False, 'message': 'New password must be at least 8 characters long'})
    
    if any(variant in new_password.lower() for variant in ['admin', 'administrator']):
        return jsonify({'success': False, 'message': 'Your password cannot contain "admin" or "administrator"'})
    
    try:
        with DBContextManager() as conn:
            admin = conn.execute('SELECT id, password_hash FROM admin WHERE id = ?', (session['admin_id'],)).fetchone()
            
            if not admin:
                logging.error(f"Admin account not found for ID: {session['admin_id']}")
                return jsonify({'success': False, 'message': 'Admin account not found. Please contact support.'})
            
            if not check_password_hash(admin['password_hash'], current_password):
                logging.warning(f"Incorrect current password provided by admin ID: {session['admin_id']}")
                return jsonify({'success': False, 'message': 'Current password is incorrect'})
            
            if current_password == new_password:
                return jsonify({'success': False, 'message': 'New password cannot be the same as current password'})
            
            new_password_hash = generate_password_hash(new_password)
            
            is_default_hash = admin['password_hash'] == '$2b$12$qKU3YP7Nz3kzWkxpYKiMqe4JfN9aKC7GW4q1Eb1iiL6TgW/LQTKCm'
            
            if is_default_hash:
                logging.info(f"Admin {session.get('admin_username')} is changing from default password")
            
            conn.execute('UPDATE admin SET password_hash = ? WHERE id = ?', (new_password_hash, session['admin_id']))
            conn.commit()
            
            if 'default_password' in session:
                session.pop('default_password', None)
                
            logging.info(f"Password successfully changed for admin {session.get('admin_username')}")
            return jsonify({'success': True, 'message': 'Password changed successfully'})
    except Exception as e:
        logging.error(f"Error in admin_change_password: {e}")
        return jsonify({'success': False, 'message': f'Error changing password: {str(e)}'})

@app.route('/admin/search_voter', methods=['POST'])
def admin_search_voter():
    if not session.get('admin'):
        return jsonify({'success': False, 'message': 'Unauthorized'}), 401
    
    query = request.form.get('query', '').strip()
    csrf_token = request.form.get('csrf_token')
    
    if not query:
        return jsonify({'success': False, 'message': 'No search query provided'})
    
    conn = get_db()
    
    try:
        if query.isdigit():
            voter = conn.execute(
                'SELECT * FROM voters WHERE id = ?',
                (query,)
            ).fetchone()
            
            if voter:
                conn.close()
                return jsonify({
                    'success': True, 
                    'voters': [dict(voter)]
                })
        
        voter = conn.execute(
            'SELECT * FROM voters WHERE voting_code = ?',
            (query,)
        ).fetchone()
        
        if voter is None:
            voter = conn.execute(
                'SELECT * FROM voters WHERE voting_code = ?',
                (query.upper(),)
            ).fetchone()
        
        if voter is None:
            voter = conn.execute(
                'SELECT * FROM voters WHERE UPPER(voting_code) = UPPER(?)',
                (query,)
            ).fetchone()
        
        if voter:
            conn.close()
            return jsonify({
                'success': True, 
                'voters': [dict(voter)]
            })
        
        voters = conn.execute(
            'SELECT * FROM voters WHERE name LIKE ? ORDER BY is_teacher, class, section, name',
            (f'%{query}%',)
        ).fetchall()
        
        conn.close()
        
        if not voters:
            return jsonify({
                'success': False, 
                'message': 'No voters found matching your search'
            })
        voters_list = [dict(v) for v in voters]
        return jsonify({
            'success': True, 
            'voters': voters_list
        })
    
    except Exception as e:
        logging.error(f"Error in search_voter: {str(e)}")
        conn.close()
        return jsonify({'success': False, 'message': f'An error occurred: {str(e)}'})

@app.route('/admin/teachers/download')
def download_all_teachers():
    if not session.get('admin'):
        return redirect(url_for('admin_login'))
    
    try:
        with DBContextManager() as conn:
            teachers = conn.execute('SELECT * FROM voters WHERE is_teacher = 1 ORDER BY name').fetchall()
            
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", size=12)
            
            pdf.cell(200, 10, txt="Pathshala Election 2082 - All Teacher Voting Codes", ln=True, align='C')
            pdf.set_font("Arial", 'B', size=14)
            pdf.cell(200, 10, txt="CONFIDENTIAL", ln=True, align='C')
            pdf.set_font("Arial", size=10)
            pdf.cell(200, 10, txt=f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align='C')
            pdf.ln(10)
            
            pdf.set_font("Arial", 'B', size=10)
            pdf.cell(10, 10, txt="ID", border=1)
            pdf.cell(80, 10, txt="Name", border=1)
            pdf.cell(50, 10, txt="Subject", border=1)
            pdf.cell(30, 10, txt="Voting Code", border=1)
            pdf.cell(30, 10, txt="Vote Status", border=1, ln=True)
            
            pdf.set_font("Arial", size=10)
            teachers_per_page = 25
            teacher_count = 0
            
            for teacher in teachers:
                pdf.cell(10, 10, txt=str(teacher['id']), border=1)
                pdf.cell(80, 10, txt=teacher['name'], border=1)
                pdf.cell(50, 10, txt=str(teacher['subject'] or ''), border=1)
                pdf.cell(30, 10, txt=teacher['voting_code'], border=1)
                vote_status = "Voted" if teacher['has_voted'] else "Not Voted"
                pdf.cell(30, 10, txt=vote_status, border=1, ln=True)
                
                teacher_count += 1
                if teacher_count % teachers_per_page == 0 and teacher_count < len(teachers):
                    pdf.add_page()
                    
                    pdf.set_font("Arial", 'B', size=10)
                    pdf.cell(10, 10, txt="ID", border=1)
                    pdf.cell(80, 10, txt="Name", border=1)
                    pdf.cell(50, 10, txt="Subject", border=1)
                    pdf.cell(30, 10, txt="Voting Code", border=1)
                    pdf.cell(30, 10, txt="Vote Status", border=1, ln=True)
                    
                    pdf.set_font("Arial", size=10)
            
            pdf_temp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
            pdf_temp.close()
            
            pdf.output(pdf_temp.name)
            
            return send_file(
                pdf_temp.name,
                mimetype='application/pdf',
                as_attachment=True,
                download_name='all_teacher_voting_codes.pdf'
            )
            
    except Exception as e:
        flash(f'Error generating PDF: {str(e)}')
        return redirect(url_for('admin_teachers'))

if __name__ == '__main__':
    if not os.path.exists(app.config['DATABASE']):
        init_db()
    
    # Production mode
    app.run(debug=False, host='0.0.0.0')